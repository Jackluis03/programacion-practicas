
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

st.title("Formulario para Programación de Prácticas")

# === Datos generales del docente ===
st.header("Datos Generales del Docente")

docente = st.text_input("Nombre del docente")
carrera = st.text_input("Carrera")
materia = st.text_input("Materia")
semestre = st.text_input("Semestre")
grupo = st.text_input("Grupo")
num_alumnos = st.number_input("Número de alumnos", min_value=1)
fecha = st.date_input("Fecha programada")
horario = st.text_input("Horario (ej. 10:00 - 12:00)")
duracion = st.text_input("Duración (ej. 2 horas)")
responsable = st.text_input("Responsable del laboratorio")

# === Lista de prácticas ===
st.header("Agregar Prácticas")
if "practicas" not in st.session_state:
    st.session_state.practicas = []

with st.form("practica_form"):
    no_practica = st.number_input("Número de práctica", min_value=1)
    nombre_practica = st.text_input("Nombre de la práctica")
    objetivo = st.text_area("Objetivo")
    materiales = st.text_area("Materiales, reactivos y equipo requerido")
    observaciones = st.text_area("Observaciones")

    submitted = st.form_submit_button("Agregar práctica")
    if submitted:
        st.session_state.practicas.append({
            "No.": no_practica,
            "Nombre de la práctica": nombre_practica,
            "Objetivo": objetivo,
            "Nombre del docente": docente,
            "Carrera": carrera,
            "Materia": materia,
            "Semestre": semestre,
            "Grupo": grupo,
            "Número de alumnos": num_alumnos,
            "Fecha programada": fecha.strftime("%d/%m/%Y"),
            "Horario": horario,
            "Duración": duracion,
            "Responsable": responsable,
            "Materiales": materiales,
            "Observaciones": observaciones
        })
        st.success("Práctica agregada correctamente.")

# === Mostrar prácticas agregadas ===
if st.session_state.practicas:
    st.subheader("Prácticas Agregadas")
    st.write(pd.DataFrame(st.session_state.practicas))

# === Guardar en Excel ===
if st.button("Exportar a Excel"):
    # Cargar plantilla existente
    plantilla_path = "PROGRAMACION DE PRACTICAS.xlsx"
    wb = load_workbook(plantilla_path)
    ws = wb.active  # asumiendo que es la hoja principal

    start_row = 8  # Donde empiezan las prácticas en la hoja

    for idx, practica in enumerate(st.session_state.practicas):
        row = start_row + idx
        ws.cell(row=row, column=1).value = practica["No."]
        ws.cell(row=row, column=2).value = practica["Nombre de la práctica"]
        ws.cell(row=row, column=3).value = practica["Objetivo"]
        ws.cell(row=row, column=4).value = practica["Nombre del docente"]
        ws.cell(row=row, column=5).value = practica["Carrera"]
        ws.cell(row=row, column=6).value = practica["Materia"]
        ws.cell(row=row, column=7).value = practica["Semestre"]
        ws.cell(row=row, column=8).value = practica["Grupo"]
        ws.cell(row=row, column=9).value = practica["Número de alumnos"]
        ws.cell(row=row, column=10).value = practica["Fecha programada"]
        ws.cell(row=row, column=11).value = practica["Horario"]
        ws.cell(row=row, column=12).value = practica["Duración"]
        ws.cell(row=row, column=13).value = practica["Responsable"]
        ws.cell(row=row, column=14).value = practica["Materiales"]
        ws.cell(row=row, column=15).value = practica["Observaciones"]

    output_file = "PRACTICAS_EXPORTADAS.xlsx"
    wb.save(output_file)
    st.success(f"Datos exportados exitosamente a {output_file}")
